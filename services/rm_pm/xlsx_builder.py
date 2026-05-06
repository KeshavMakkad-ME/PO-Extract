import io
import logging
import os
from copy import deepcopy

import openpyxl
import pandas as pd

from config.rm_pm.field_map import NUMBER_FIELDS

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "templates/Finance_XLSX_Templates_new_new.xlsx")
SHEET_NAME = "Rm And PM"
DATA_START_ROW = 2


def _col_to_idx(col: str) -> int:
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - 64)
    return result


def _coerce_number(value, field_name: str):
    if field_name in NUMBER_FIELDS and value not in ("", None):
        try:
            v = float(str(value).replace(",", ""))
            return int(v) if v == int(v) else v
        except (ValueError, TypeError):
            pass
    return value


def _to_num(value) -> float:
    if value in (None, "", "None"):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _clear_data_rows(ws, config_df: pd.DataFrame) -> None:
    managed_cols = {_col_to_idx(cfg["col_letter"]) for _, cfg in config_df.iterrows()}
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        for cell in row:
            if cell.column in managed_cols:
                cell.value = None
    logger.info(f"Cleared existing data rows in '{SHEET_NAME}'")


def build_xlsx(results: list, config_df: pd.DataFrame, freight_df: pd.DataFrame) -> tuple[bytes, int]:
    """
    Load the template, clear existing data, write extracted invoice data into
    'Rm And PM' starting at row 2, and return the file as bytes.

    invoice_amount is computed per batch row (taxable + tax amounts).
    freight_charges is written as a dedicated row after all batch rows.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    _clear_data_rows(ws, config_df)

    # field_name -> column index (used for derived calcs and freight row)
    field_to_col = {
        cfg["field_name"]: _col_to_idx(cfg["col_letter"])
        for _, cfg in config_df.iterrows()
    }

    data_rows: list[dict] = []

    for invoice_result in results:
        invoice_result = deepcopy(invoice_result)

        if "_error" in invoice_result:
            logger.warning(f"Skipping invoice: {invoice_result['_error']}")
            continue

        # Derived invoice-level fields — never from LLM
        invoice_result["voucher_number"] = invoice_result.get("supplier_invoice_number", "")
        supplier_no = invoice_result.get("supplier_invoice_number", "Unknown")
        party = invoice_result.get("party_name", "Unknown")
        invoice_result["narration"] = f"Invoice {supplier_no} from {party}"

        line_items = invoice_result.get("line_items", [])
        if not line_items:
            logger.warning(f"Invoice {supplier_no} has no line items — skipping")
            continue

        for item in line_items:
            row_data: dict[int, object] = {}
            item_values: dict[str, object] = {}

            for _, cfg in config_df.iterrows():
                col_idx    = _col_to_idx(cfg["col_letter"])
                field_name = cfg["field_name"]
                source     = cfg["source"]

                if source == "hardcoded":
                    value = cfg["hardcoded_value"]
                elif field_name == "invoice_amount":
                    value = ""  # always derived below — never from LLM or config
                elif field_name == "freight_charges":
                    value = ""  # always written as its own row — not on batch rows
                elif source == "extracted":
                    value = item.get(field_name) or invoice_result.get(field_name, "")
                elif source == "derived":
                    value = invoice_result.get(field_name, "")
                else:
                    value = ""

                coerced = _coerce_number(value or "", field_name)
                final_val = coerced if coerced != "" else None
                row_data[col_idx] = final_val
                item_values[field_name] = final_val

            # Compute invoice_amount = taxable_amount + igst_amount + cgst_amount + sgst_amount
            if "invoice_amount" in field_to_col:
                inv_amt = (
                    _to_num(item_values.get("taxable_amount"))
                    + _to_num(item_values.get("igst_amount"))
                    + _to_num(item_values.get("cgst_amount"))
                    + _to_num(item_values.get("sgst_amount"))
                )
                if inv_amt:
                    row_data[field_to_col["invoice_amount"]] = (
                        int(inv_amt) if inv_amt == int(inv_amt) else round(inv_amt, 2)
                    )

            data_rows.append(row_data)

        # Dedicated freight row after all batch rows for this invoice
        freight = _to_num(invoice_result.get("freight_charges"))

        if freight:
            freight_row: dict[int, object] = {}

            # Fill all fields using same logic as normal rows
            for _, cfg in config_df.iterrows():
                col_idx = _col_to_idx(cfg["col_letter"])
                field_name = cfg["field_name"]
                source = cfg["source"]

                if source == "hardcoded":
                    value = cfg["hardcoded_value"]

                elif field_name == "invoice_amount":
                    value = freight

                elif field_name == "freight_charges":
                    value = freight

                elif source == "extracted":
                    value = invoice_result.get(field_name, "")

                elif source == "derived":
                    value = invoice_result.get(field_name, "")

                else:
                    value = ""

                coerced = _coerce_number(value or "", field_name)
                final_val = coerced if coerced != "" else None

                freight_row[col_idx] = final_val

            # Override narration specifically for freight row
            if "narration" in field_to_col:
                freight_row[field_to_col["narration"]] = "Freight Charges"

            data_rows.append(freight_row)

            logger.info(f"Invoice {supplier_no} — added freight row ({freight})")

    for r_idx, row_data in enumerate(data_rows, start=DATA_START_ROW):
        for col_idx, value in row_data.items():
            ws.cell(row=r_idx, column=col_idx, value=value)

    if data_rows:
        logger.info(f"Wrote {len(data_rows)} rows into '{SHEET_NAME}'")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), len(data_rows)
