import io
import logging
import os
from copy import deepcopy

import openpyxl
import pandas as pd

from config.po.field_map import NUMBER_FIELDS

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "templates/Finance_XLSX_Templates_new_new.xlsx")

# Maps field_config field_name → key in the dispatch-from JSON object
_DISPATCH_KEY_MAP = {
    "dispatch_from_name":     "name",
    "dispatch_from_location": "Location",
    "dispatch_from_pincode":  "Pincode",
    "dispatch_from_state":    "State",
    "dispatch_from_address1": "Address",
}


def _col_to_idx(col: str) -> int:
    """Column letter to 1-based index. 'A'->1, 'Z'->26, 'AA'->27, 'DJ'->114."""
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - 64)
    return result


def _get_state_from_gstin(gstin: str, state_codes: dict) -> str:
    if not gstin or len(gstin) < 2:
        logger.warning(f"Invalid GSTIN: '{gstin}' — cannot derive state")
        return ""
    prefix = gstin[:2]
    for _, code in state_codes.items():
        if code.startswith(f"{prefix}-"):
            return code
    logger.warning(f"Unknown state prefix '{prefix}' from GSTIN '{gstin}'")
    return ""


def _get_address2(address1: str) -> str:
    if not address1:
        return ""
    parts = [p.strip() for p in address1.split(",") if p.strip()]
    return parts[-1] if parts else ""


def _coerce_number(value, field_name: str):
    if field_name in NUMBER_FIELDS and value not in ("", None):
        try:
            v = float(value)
            return int(v) if v == int(v) else v
        except (ValueError, TypeError):
            pass
    return value


def _clear_data_rows(ws, config_df: pd.DataFrame) -> None:
    """Clear values in managed columns from row 4 downwards, preserving formatting."""
    if ws.max_row < 4:
        return
    managed_cols = {_col_to_idx(cfg["col_letter"]) for _, cfg in config_df.iterrows()}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            if cell.column in managed_cols:
                cell.value = None
    logger.info(f"Cleared rows 4–{ws.max_row} in 'E Invoice' sheet")


def build_xlsx(
    results: list,
    config_df: pd.DataFrame,
    state_codes: dict,
    dispatch_from: dict | None = None,
) -> tuple[bytes, int]:
    """
    Loads the local XLSX template, clears existing data rows, writes extracted PO data
    into 'E Invoice' starting at row 4, and returns the modified file as bytes.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["E Invoice"]

    _clear_data_rows(ws, config_df)

    dispatch_from = dispatch_from or {}
    data_rows: list[dict] = []

    for po_result in results:
        po_result = deepcopy(po_result)

        if "_error" in po_result:
            logger.warning(f"Skipping PO: {po_result['_error']}")
            continue

        # Derived fields — always computed here, never from the LLM
        po_result["party_state"]        = _get_state_from_gstin(po_result.get("party_gstin", ""), state_codes)
        po_result["ship_to_gstin"]      = po_result.get("party_gstin", "")
        po_result["ship_to_trade_name"] = po_result.get("party_trade_name", "")
        po_result["ship_to_state"]      = po_result.get("party_state", "")
        po_result["party_address2"]     = _get_address2(po_result.get("party_address1", ""))
        po_result["ship_to_address2"]   = _get_address2(po_result.get("ship_to_address1", ""))

        line_items = po_result.get("line_items", [])
        if not line_items:
            logger.warning(f"PO {po_result.get('invoice_number')} has no line items — skipping")
            continue

        for item in line_items:
            row_data: dict[int, object] = {}

            for _, cfg in config_df.iterrows():
                col_idx    = _col_to_idx(cfg["col_letter"])  # 1-based
                field_name = cfg["field_name"]
                source     = cfg["source"]

                if field_name == "dispatch_from_address2":
                    # always derived from dispatch address regardless of source in config
                    value = _get_address2(dispatch_from.get("Address", ""))
                elif source == "hardcoded":
                    value = cfg["hardcoded_value"]
                elif source == "extracted":
                    value = item.get(field_name) or po_result.get(field_name, "")
                elif source == "derived":
                    value = po_result.get(field_name, "")
                elif source == "dispatch_from":
                    if field_name == "dispatch_from_state":
                        raw = dispatch_from.get("State", "")
                        value = state_codes.get(raw.lower(), raw) if raw else ""
                    else:
                        json_key = _DISPATCH_KEY_MAP.get(field_name, field_name)
                        value = dispatch_from.get(json_key, "")
                else:
                    value = ""

                coerced = _coerce_number(value or "", field_name)
                row_data[col_idx] = coerced if coerced != "" else None

            data_rows.append(row_data)

    for r_idx, row_data in enumerate(data_rows, start=4):
        for col_idx, value in row_data.items():
            ws.cell(row=r_idx, column=col_idx, value=value)

    if data_rows:
        logger.info(f"Wrote {len(data_rows)} rows into template 'E Invoice' sheet")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), len(data_rows)
