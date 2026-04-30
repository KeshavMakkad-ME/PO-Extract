import json
import logging
import os

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "templates/Finance_XLSX_Templates_new_new.xlsx")


def load_field_config() -> pd.DataFrame:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    ws = wb["field_config"]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = [[str(v) if v is not None else "" for v in row] for row in rows[1:]]
    wb.close()

    df = pd.DataFrame(data, columns=headers)
    df = df[df["field_name"].notna() & (df["field_name"] != "")].reset_index(drop=True)
    logger.info(f"Loaded field_config: {len(df)} fields from template")
    return df


def load_state_codes() -> dict:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    ws = wb["Drop down -- Keshav "]

    state_codes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 3 and row[3]:
            full_code = str(row[3]).strip()
            if "-" in full_code:
                state_name = full_code.split("-", 1)[1].strip().lower()
                state_codes[state_name] = full_code

    wb.close()
    logger.info(f"Loaded {len(state_codes)} state codes from template")
    return state_codes


def load_dispatch_options() -> list[dict]:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    ws = wb["Drop down -- Keshav "]

    options = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 10 and row[10]:
            try:
                options.append(json.loads(str(row[10])))
            except (json.JSONDecodeError, TypeError):
                pass

    wb.close()
    logger.info(f"Loaded {len(options)} dispatch options from template")
    return options
