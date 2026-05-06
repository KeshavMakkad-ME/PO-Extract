import logging
import os

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "templates/Finance_XLSX_Templates_new_new.xlsx")


def load_field_config() -> pd.DataFrame:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    ws = wb["field_config_rm_pm"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) for h in rows[0]]
    data = [[str(v) if v is not None else "" for v in row] for row in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    df = df[df["field_name"].notna() & (df["field_name"] != "")].reset_index(drop=True)
    logger.info(f"Loaded field_config_rm_pm: {len(df)} fields")
    return df


def load_rm_pm_options() -> dict:
    """Load voucher type names (col L) and purchase ledgers (col M) from the dropdown sheet."""
    wb  = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    ws  = wb["Drop down -- Keshav "]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    voucher_type_names = [str(r[11]) for r in rows if len(r) > 11 and r[11] is not None]
    purchase_ledgers   = [str(r[12]) for r in rows if len(r) > 12 and r[12] is not None]

    logger.info(
        f"Loaded RM/PM options — {len(voucher_type_names)} voucher types, "
        f"{len(purchase_ledgers)} purchase ledgers"
    )
    return {"voucher_type_names": voucher_type_names, "purchase_ledgers": purchase_ledgers}
